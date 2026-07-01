from __future__ import annotations

import csv
import io
import json
import math
import os
import re
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from flask import Flask, jsonify, redirect, render_template, request, send_file, url_for
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Flask belum terinstall. Jalankan: python3 -m pip install -r requirements.txt"
    ) from exc


BASE_DIR = Path(__file__).resolve().parent
MODEL_DIR = BASE_DIR / "model_mdeberta_zero_shot"
UPLOAD_DIR = BASE_DIR / "uploads"
EXPORT_DIR = BASE_DIR / "exports"
LABEL_CONFIG = MODEL_DIR / "label_config.json"

UPLOAD_DIR.mkdir(exist_ok=True)
EXPORT_DIR.mkdir(exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024


DEFAULT_LABELS = [
    "Kualitas Barang",
    "Keamanan Kemasan",
    "Kecepatan Pengiriman",
    "Pelayanan Toko",
]

SENTIMENT_LABELS = ["Positif", "Netral", "Negatif"]

LABEL_META = {
    "Kualitas Barang": {
        "color": "#22c55e",
        "soft": "#dcfce7",
        "icon": "star",
        "description": "Terkait kualitas & kesesuaian produk",
        "keywords": ["bagus", "original", "kualitas", "produk", "sesuai", "fungsi", "harga", "rusak"],
    },
    "Keamanan Kemasan": {
        "color": "#f97316",
        "soft": "#ffedd5",
        "icon": "package",
        "description": "Terkait kondisi & keamanan packaging",
        "keywords": ["packing", "kemasan", "aman", "rapi", "penyok", "bungkus", "pecah", "box"],
    },
    "Kecepatan Pengiriman": {
        "color": "#0ea5e9",
        "soft": "#e0f2fe",
        "icon": "truck",
        "description": "Terkait waktu & ketepatan pengiriman",
        "keywords": ["kirim", "pengiriman", "cepat", "lama", "kurir", "datang", "sampai", "terlambat"],
    },
    "Pelayanan Toko": {
        "color": "#7c3aed",
        "soft": "#ede9fe",
        "icon": "messages-square",
        "description": "Terkait respon & pelayanan seller",
        "keywords": ["seller", "toko", "respon", "ramah", "pelayanan", "membantu", "chat", "balas"],
    },
}

SENTIMENT_META = {
    "Positif": {
        "color": "#16a34a",
        "soft": "#dcfce7",
        "icon": "check-circle",
        "keywords": ["bagus", "cepat", "rapi", "aman", "original", "sesuai", "ramah", "membantu", "baik", "puas"],
    },
    "Netral": {
        "color": "#64748b",
        "soft": "#f1f5f9",
        "icon": "info",
        "keywords": ["cukup", "standar", "biasa", "lumayan", "oke", "sesuai", "diterima", "sampai"],
    },
    "Negatif": {
        "color": "#dc2626",
        "soft": "#fee2e2",
        "icon": "alert-circle",
        "keywords": ["lama", "rusak", "penyok", "kurang", "tidak", "buruk", "kecewa", "lambat", "pecah", "hilang"],
    },
}

STOPWORDS = {
    "dan", "yang", "di", "ke", "dari", "ini", "itu", "dengan", "untuk", "atau", "pada",
    "sangat", "sudah", "tidak", "nya", "aja", "ada", "dalam", "barang", "produk",
}

APP_STATE: dict[str, Any] = {
    "dataset": [],
    "filename": "",
    "uploaded_at": None,
    "results": [],
    "summary": {},
}

_MODEL_PIPELINE = None
_MODEL_ERROR = None


def load_label_config() -> tuple[list[str], str]:
    if LABEL_CONFIG.exists():
        data = json.loads(LABEL_CONFIG.read_text(encoding="utf-8"))
        return data.get("candidate_labels", DEFAULT_LABELS), data.get(
            "hypothesis_template", "Ulasan ini berfokus secara tajam pada topik {}."
        )
    return DEFAULT_LABELS, "Ulasan ini berfokus secara tajam pada topik {}."


CANDIDATE_LABELS, HYPOTHESIS_TEMPLATE = load_label_config()


def clean_text(text: str) -> str:
    text = str(text).lower()
    text = re.sub(r"http\S+|www\.\S+", " ", text)
    text = re.sub(r"[^a-zA-Z0-9\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def word_count(text: str) -> int:
    return len(clean_text(text).split())


def get_model_pipeline():
    global _MODEL_PIPELINE, _MODEL_ERROR
    if os.environ.get("INTENTCLASSIFY_USE_MODEL", "1").lower() in {"0", "false", "no"}:
        _MODEL_ERROR = "Mode demo cepat aktif: INTENTCLASSIFY_USE_MODEL=0."
        return None
    if _MODEL_PIPELINE is not None or _MODEL_ERROR is not None:
        return _MODEL_PIPELINE
    try:
        import torch
        from transformers import pipeline

        device = 0 if torch.cuda.is_available() else -1
        _MODEL_PIPELINE = pipeline(
            "zero-shot-classification",
            model=str(MODEL_DIR),
            tokenizer=str(MODEL_DIR),
            device=device,
        )
    except Exception as exc:  # pragma: no cover - depends on local ML stack
        _MODEL_ERROR = str(exc)
        _MODEL_PIPELINE = None
    return _MODEL_PIPELINE


def keyword_predict(text: str) -> dict[str, Any]:
    clean = clean_text(text)
    scores: dict[str, float] = {}
    for label in CANDIDATE_LABELS:
        keywords = LABEL_META[label]["keywords"]
        hit_count = sum(1 for keyword in keywords if keyword in clean)
        scores[label] = hit_count + 0.2

    total = sum(scores.values()) or 1
    ranked = sorted(
        [{"label": label, "score": score / total} for label, score in scores.items()],
        key=lambda item: item["score"],
        reverse=True,
    )
    if ranked[0]["score"] < 0.45:
        ranked[0]["score"] = min(0.76, ranked[0]["score"] + 0.25)
    return {"label": ranked[0]["label"], "score": ranked[0]["score"], "ranking": ranked}


def keyword_sentiment_predict(text: str) -> dict[str, Any]:
    clean = clean_text(text)
    scores: dict[str, float] = {}
    for label in SENTIMENT_LABELS:
        keywords = SENTIMENT_META[label]["keywords"]
        hit_count = sum(1 for keyword in keywords if keyword in clean)
        scores[label] = hit_count + 0.2

    total = sum(scores.values()) or 1
    ranked = sorted(
        [{"label": label, "score": score / total} for label, score in scores.items()],
        key=lambda item: item["score"],
        reverse=True,
    )
    if ranked[0]["score"] < 0.45:
        ranked[0]["score"] = min(0.72, ranked[0]["score"] + 0.2)
    return {"label": ranked[0]["label"], "score": ranked[0]["score"], "ranking": ranked}


def zero_shot_classify(text: str, labels: list[str], hypothesis_template: str) -> dict[str, Any] | None:
    classifier = get_model_pipeline()
    if classifier:
        result = classifier(
            text,
            labels,
            hypothesis_template=hypothesis_template,
        )
        ranking = [
            {"label": label, "score": float(score)}
            for label, score in zip(result["labels"], result["scores"])
        ]
        return {"label": result["labels"][0], "score": float(result["scores"][0]), "ranking": ranking}
    return None


def classify_review(text: str) -> dict[str, Any]:
    prediction = zero_shot_classify(text, CANDIDATE_LABELS, HYPOTHESIS_TEMPLATE)
    if prediction:
        return prediction
    return keyword_predict(text)


def classify_sentiment(text: str) -> dict[str, Any]:
    prediction = zero_shot_classify(
        text,
        SENTIMENT_LABELS,
        "Sentimen ulasan pelanggan ini adalah {}.",
    )
    if prediction:
        return prediction
    return keyword_sentiment_predict(text)


def read_csv_file(path: Path) -> list[dict[str, Any]]:
    raw = path.read_bytes()
    text = raw.decode("utf-8-sig", errors="replace")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    return [dict(row) for row in reader]


def read_xlsx_file(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("File XLSX membutuhkan openpyxl. Install requirements.txt terlebih dahulu.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or "").strip() for cell in rows[0]]
    data = []
    for row in rows[1:]:
        data.append({headers[i] or f"kolom_{i+1}": row[i] if i < len(row) else "" for i in range(len(headers))})
    return data


def find_review_column(rows: list[dict[str, Any]]) -> str | None:
    if not rows:
        return None
    columns = list(rows[0].keys())
    preferred = ["review", "review_text", "ulasan", "komentar", "text", "teks", "content"]
    lowered = {col.lower().strip(): col for col in columns}
    for name in preferred:
        if name in lowered:
            return lowered[name]
    return columns[0] if columns else None


def normalize_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    review_col = find_review_column(rows)
    normalized = []
    for row in rows:
        review = str(row.get(review_col, "") if review_col else "").strip()
        normalized.append({"review": review, "length": word_count(review), "raw": row})
    return normalized


def length_distribution(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets = [("1-5", 1, 5), ("6-10", 6, 10), ("11-15", 11, 15), ("16-20", 16, 20), ("21-25", 21, 25), ("26-30", 26, 30), (">30", 31, 10**9)]
    counts = []
    for label, low, high in buckets:
        counts.append({"label": label, "count": sum(1 for row in rows if low <= row["length"] <= high)})
    return counts


def common_words(rows: list[dict[str, Any]], limit: int = 26) -> list[dict[str, Any]]:
    counter: Counter[str] = Counter()
    for row in rows:
        counter.update(word for word in clean_text(row["review"]).split() if len(word) > 2 and word not in STOPWORDS)
    top = counter.most_common(limit)
    if not top:
        return []
    max_count = top[0][1]
    return [{"word": word, "weight": 12 + int((count / max_count) * 30)} for word, count in top]


def classify_dataset(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    results = []
    for index, row in enumerate(rows, start=1):
        pred = classify_review(row["review"])
        sentiment = classify_sentiment(row["review"])
        results.append(
            {
                "no": index,
                "review": row["review"],
                "length": row["length"],
                "predicted_label": pred["label"],
                "confidence": round(pred["score"] * 100, 2),
                "ranking": pred["ranking"],
                "sentiment": sentiment["label"],
                "sentiment_confidence": round(sentiment["score"] * 100, 2),
                "sentiment_ranking": sentiment["ranking"],
            }
        )
    return results


def build_summary(rows: list[dict[str, Any]], results: list[dict[str, Any]], elapsed: float = 0) -> dict[str, Any]:
    total = len(rows)
    category_counts = Counter(item["predicted_label"] for item in results)
    sentiment_counts = Counter(item["sentiment"] for item in results)
    distribution = []
    for label in CANDIDATE_LABELS:
        count = category_counts.get(label, 0)
        distribution.append(
            {
                "label": label,
                "count": count,
                "percent": round((count / total * 100), 1) if total else 0,
                "color": LABEL_META[label]["color"],
                "keywords": LABEL_META[label]["keywords"][:4],
            }
        )
    sentiment_distribution = []
    for label in SENTIMENT_LABELS:
        count = sentiment_counts.get(label, 0)
        sentiment_distribution.append(
            {
                "label": label,
                "count": count,
                "percent": round((count / total * 100), 1) if total else 0,
                "color": SENTIMENT_META[label]["color"],
                "soft": SENTIMENT_META[label]["soft"],
                "icon": SENTIMENT_META[label]["icon"],
            }
        )
    avg_length = round(sum(row["length"] for row in rows) / total) if total else 0
    avg_confidence = round(sum(item["confidence"] for item in results) / len(results), 2) if results else 0
    avg_sentiment_confidence = round(sum(item["sentiment_confidence"] for item in results) / len(results), 2) if results else 0
    return {
        "total": total,
        "categorized": len(results),
        "category_count": len([item for item in distribution if item["count"] > 0]) or len(CANDIDATE_LABELS),
        "avg_length": avg_length,
        "avg_confidence": avg_confidence,
        "highest_confidence": max([item["confidence"] for item in results], default=0),
        "lowest_confidence": min([item["confidence"] for item in results], default=0),
        "low_confidence_count": sum(1 for item in results if item["confidence"] < 50),
        "avg_sentiment_confidence": avg_sentiment_confidence,
        "elapsed": elapsed,
        "distribution": distribution,
        "sentiment_distribution": sentiment_distribution,
        "length_distribution": length_distribution(rows),
        "words": common_words(rows),
        "top_category": max(distribution, key=lambda item: item["count"], default=None),
        "top_sentiment": max(sentiment_distribution, key=lambda item: item["count"], default=None),
    }


def format_elapsed(seconds: float) -> str:
    seconds = int(seconds)
    return f"00:{seconds // 60:02d}:{seconds % 60:02d}"


def paginate(items: list[dict[str, Any]], page: int, per_page: int = 10) -> dict[str, Any]:
    total_pages = max(1, math.ceil(len(items) / per_page))
    page = min(max(page, 1), total_pages)
    start = (page - 1) * per_page
    return {
        "rows": items[start : start + per_page],
        "page": page,
        "per_page": per_page,
        "total": len(items),
        "total_pages": total_pages,
        "start": start + 1 if items else 0,
        "end": min(start + per_page, len(items)),
    }


def ensure_demo_data():
    if APP_STATE["dataset"]:
        return
    demo = [
        "Barang bagus dan original, sesuai deskripsi.",
        "Pengiriman sangat lama, padahal sudah ekspres.",
        "Packing kurang aman, barang sampai dalam keadaan rusak.",
        "Seller ramah dan fast response.",
        "Kualitas produk bagus, harga terjangkau.",
        "Barang diterima dengan cepat, pengemasan rapi.",
        "Kurir lambat dan tidak update status pengiriman.",
        "Produk sesuai foto, berfungsi dengan baik.",
        "Kotak penyok, tapi isinya aman.",
        "Pelayanan toko sangat membantu.",
    ]
    rows = [{"review": text, "length": word_count(text), "raw": {"review": text}} for text in demo]
    start = time.time()
    results = classify_dataset(rows)
    APP_STATE.update(
        {
            "dataset": rows,
            "filename": "sample_reviews.csv",
            "uploaded_at": datetime.now(),
            "results": results,
            "summary": build_summary(rows, results, time.time() - start),
        }
    )


@app.context_processor
def inject_globals():
    return {
        "labels": CANDIDATE_LABELS,
        "meta": LABEL_META,
        "sentiment_labels": SENTIMENT_LABELS,
        "sentiment_meta": SENTIMENT_META,
        "model_error": _MODEL_ERROR,
        "model_name": "mDeBERTa-v3 Zero-Shot",
        "model_version": "v1.0",
        "format_elapsed": format_elapsed,
    }


@app.route("/")
def input_page():
    return render_template("visualisasi.html", page="input")


@app.post("/api/analyze")
def api_analyze():
    payload = request.get_json(silent=True) or {}
    text = str(payload.get("text", "")).strip()
    if not text:
        return jsonify({"error": "Masukkan ulasan terlebih dahulu."}), 400
    prediction = classify_review(text)
    sentiment = classify_sentiment(text)
    return jsonify(
        {
            "review": text,
            "label": prediction["label"],
            "confidence": round(prediction["score"] * 100, 2),
            "ranking": [
                {
                    "label": item["label"],
                    "confidence": round(item["score"] * 100, 2),
                    "color": LABEL_META[item["label"]]["color"],
                }
                for item in prediction["ranking"]
            ],
            "sentiment": {
                "label": sentiment["label"],
                "confidence": round(sentiment["score"] * 100, 2),
                "ranking": [
                    {
                        "label": item["label"],
                        "confidence": round(item["score"] * 100, 2),
                        "color": SENTIMENT_META[item["label"]]["color"],
                    }
                    for item in sentiment["ranking"]
                ],
            },
        }
    )


@app.route("/dataset", methods=["GET", "POST"])
def dataset_upload():
    error = None
    if request.method == "POST":
        file = request.files.get("dataset")
        if not file or not file.filename:
            error = "Pilih file CSV atau XLSX terlebih dahulu."
        else:
            suffix = Path(file.filename).suffix.lower()
            if suffix not in {".csv", ".xlsx"}:
                error = "Format file harus CSV atau XLSX."
            else:
                saved_path = UPLOAD_DIR / f"{int(time.time())}_{Path(file.filename).name}"
                file.save(saved_path)
                try:
                    raw_rows = read_xlsx_file(saved_path) if suffix == ".xlsx" else read_csv_file(saved_path)
                    rows = normalize_rows(raw_rows)
                    start = time.time()
                    results = classify_dataset(rows)
                    APP_STATE.update(
                        {
                            "dataset": rows,
                            "filename": Path(file.filename).name,
                            "uploaded_at": datetime.now(),
                            "results": results,
                            "summary": build_summary(rows, results, time.time() - start),
                        }
                    )
                    return redirect(url_for("dataset_preview"))
                except Exception as exc:
                    error = str(exc)
    return render_template("visualisasi.html", page="upload", step=1, error=error)


@app.route("/dataset/preview")
def dataset_preview():
    ensure_demo_data()
    page_num = int(request.args.get("page", 1))
    data = paginate(APP_STATE["dataset"], page_num)
    return render_template("visualisasi.html", page="preview", step=2, data=data, summary=APP_STATE["summary"])


@app.route("/dataset/summary")
def dataset_summary():
    ensure_demo_data()
    return render_template("visualisasi.html", page="summary", step=3, summary=APP_STATE["summary"])


@app.route("/dataset/results")
def dataset_results():
    ensure_demo_data()
    page_num = int(request.args.get("page", 1))
    query = request.args.get("q", "").strip().lower()
    results = APP_STATE["results"]
    if query:
        results = [
            item
            for item in results
            if query in item["review"].lower()
            or query in item["predicted_label"].lower()
            or query in item["sentiment"].lower()
        ]
    data = paginate(results, page_num)
    return render_template("visualisasi.html", page="results", step=4, data=data, summary=APP_STATE["summary"], query=query)


@app.route("/dataset/download")
def dataset_download():
    ensure_demo_data()
    return render_template("visualisasi.html", page="download", step=5, summary=APP_STATE["summary"])


@app.route("/about")
def about():
    return render_template("visualisasi.html", page="about")


def export_csv() -> io.BytesIO:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["review", "predicted_label", "confidence_score", "sentiment", "sentiment_confidence_score", "length_words"])
    for item in APP_STATE["results"]:
        writer.writerow([item["review"], item["predicted_label"], item["confidence"], item["sentiment"], item["sentiment_confidence"], item["length"]])
    return io.BytesIO(output.getvalue().encode("utf-8-sig"))


def export_pdf() -> io.BytesIO:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except ImportError:
        text = "ReportLab belum terinstall. Install requirements.txt untuk export PDF."
        return io.BytesIO(text.encode("utf-8"))
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 48
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(48, y, "Laporan Klasifikasi Intent")
    y -= 28
    pdf.setFont("Helvetica", 10)
    summary = APP_STATE["summary"]
    lines = [
        f"Total data: {summary['total']}",
        f"Rata-rata confidence: {summary['avg_confidence']}%",
        f"Rata-rata confidence sentimen: {summary['avg_sentiment_confidence']}%",
        f"Rata-rata panjang ulasan: {summary['avg_length']} kata",
        "",
        "Distribusi kategori:",
    ]
    for item in summary["distribution"]:
        lines.append(f"- {item['label']}: {item['count']} ({item['percent']}%)")
    lines.append("")
    lines.append("Distribusi sentimen:")
    for item in summary["sentiment_distribution"]:
        lines.append(f"- {item['label']}: {item['count']} ({item['percent']}%)")
    for line in lines:
        pdf.drawString(48, y, line)
        y -= 18
    pdf.save()
    buffer.seek(0)
    return buffer


@app.route("/download/<filetype>")
def download_file(filetype: str):
    ensure_demo_data()
    if filetype == "csv":
        return send_file(export_csv(), as_attachment=True, download_name="hasil_klasifikasi.csv", mimetype="text/csv")
    if filetype == "xlsx":
        return send_file(export_csv(), as_attachment=True, download_name="hasil_klasifikasi.xlsx", mimetype="text/csv")
    if filetype == "pdf":
        return send_file(export_pdf(), as_attachment=True, download_name="laporan_klasifikasi.pdf", mimetype="application/pdf")
    return redirect(url_for("dataset_download"))


if __name__ == "__main__":
    app.run(debug=True)
